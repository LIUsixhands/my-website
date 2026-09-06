# -*- coding: utf-8 -*-
"""產出／併入「員林交流道特定區 A區 地主同意書整合名冊（出勤版）」

用法:
  python3 build_roster.py data.json out.xlsx [區塊] [--field 現場名單.xlsx]

- data.json  ：parse_deed.py 解析出的謄本資料（權利、面積以此為準）
- --field    ：現場跑地主名單（全名／電話／地址／意願等級／拜訪摘要／承辦人），
               僅首次或需要重新灌入時使用；之後人工填寫以 out.xlsx 為準。
可重複執行：既有 Excel 的人工欄位與拜訪紀錄會原樣保留。
"""
import json, os, re, sys, datetime, unicodedata
from fractions import Fraction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import fieldload as FL

PING = 3.305785
ZH_DIGIT = str.maketrans('０１２３４５６７８９', '0123456789')

WILL = '"A,B,C,D,E,未接觸"'
STAGE = '"未接觸,已電聯,約訪中,已面談,待回覆,口頭同意,已送同意書,已收回同意書,婉拒,聯絡不上"'
SEAL = '"未備,已備,不需"'
KEY = '"邊緣,一般,卡中央,卡道路,卡出入口"'

# ── 人工欄位（併檔時保留）
M_OWNER = ['全名', '電話', '通訊地址', '意願等級', '拜訪階段', '承辦人', '關鍵人註記',
           '最近接觸日', '下次追蹤日', '同意書狀態', '紙本收件日', '印鑑證明', '最新拜訪摘要', '備註']
M_LOT = ['同意書狀態', '紙本收件日', '印鑑證明', '聯絡方式', '上次接觸日', '下次追蹤日',
         '位置關鍵性', '備註']

H_OWNER = ['優先序', '地主代號', '全名', '謄本姓名', '統一編號', '電話', '通訊地址',
           '所有權人類型', '持有地號', '登記筆數', '合計持分面積(㎡)', '合計持分面積(坪)',
           '佔已建檔比例', '意願等級', '拜訪階段', '承辦人', '關鍵人註記', '最近接觸日',
           '下次追蹤日', '同意書狀態', '紙本收件日', '印鑑證明', '最新拜訪摘要', '備註', '地主鍵值']

H_LOT = ['區塊', '鄉鎮市', '段別', '地號', '地號全碼', '地目', '宗地面積(㎡)', '宗地面積(坪)',
         '公告現值(元/㎡)', '登記次序', '地主代號', '全名', '謄本姓名', '統一編號',
         '所有權人類型', '管理者', '權利範圍', '持分型態', '持分比例', '持分面積(㎡)',
         '持分面積(坪)', '持分現值(元)', '電話', '通訊地址', '意願等級', '承辦人',
         '登記日期', '登記原因', '權狀字號', '設定他項權利', '其他登記事項'] + M_LOT + \
        ['謄本列印時間', '來源檔案', '資料鍵值']

H_SUM = ['區塊', '段別', '地號', '地號全碼', '地目', '宗地面積(㎡)', '宗地面積(坪)',
         '公告現值(元/㎡)', '所有權人數', '持分型態', '持分合計檢核', '公私有別',
         '他項權利筆數', '意願A人數', '同意進度', '位置關鍵性', '備註']

H_VISIT = ['日期', '地主代號', '全名', '謄本姓名', '地號', '承辦人', '接觸方式', '對象',
           '內容摘要', '意願等級', '下次行動', '下次日期', '登錄者']

H_RIGHT = ['區塊', '地號', '地號全碼', '序號', '登記次序', '權利種類', '權利人', '權利人統一編號',
           '擔保債權總金額', '標的登記次序', '設定權利範圍', '存續期間', '共同擔保地號',
           '登記日期', '登記原因']

H_TODO = ['類別', '地號', '對象', '統一編號', '說明', '建議處理', '負責人', '完成日']

H_FILL = PatternFill('solid', fgColor='1F3864')
M_FILL = PatternFill('solid', fgColor='C55A11')
T_FILL = PatternFill('solid', fgColor='7B3F00')
THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)


def share_type(rv):
    rv = rv or ''
    if '公同共有' in rv: return '公同共有'
    m = re.search(r'(\d+)分之(\d+)', rv)
    if m: return '單獨所有' if m.group(1) == m.group(2) else '分別共有'
    return '待確認'


def frac(s):
    if not s: return None
    s = s.replace('全部', '').replace('公同共有', '').strip()
    m = re.search(r'(\d+)分之(\d+)', s)
    return Fraction(int(m.group(2)), int(m.group(1))) if m else None


def joint_groups(owners):
    gid, members = {}, {}
    for o in owners:
        if '公同共有' not in (o.get('權利範圍') or ''): continue
        seq = o.get('登記次序', '')
        note = (o.get('其他登記事項') or '').translate(ZH_DIGIT)
        m = re.search(r'主登記([\d．\.、,\s]+)為公同共有', note)
        key = ('note', tuple(sorted(int(x) for x in re.findall(r'\d+', m.group(1))))) if m \
            else ('rv', o.get('權利範圍'))
        gid[seq] = key
        members.setdefault(key, set()).add(seq)
    gsize = {k: (len(k[1]) if k[0] == 'note' else len(v)) for k, v in members.items()}
    return gid, gsize


def owner_type(name, uid):
    n, u = name or '', uid or ''
    if n == '中華民國' or '國有' in n: return '公有(國有)'
    if re.search(r'(縣|市|鄉|鎮)政府$', n): return '公有(地方)'
    if re.search(r'(公司|農會|銀行|合作社|工廠|企業社|寺|宮|廟|協會|基金會|學校)', n): return '法人/團體'
    if re.fullmatch(r'\d{8}', u): return '法人/團體'
    if re.fullmatch(r'\d{10}', u): return '公有(國有)'
    if re.search(r'[A-Za-z]', u): return '自然人'
    return '待確認'


def norm_addr(a):
    a = unicodedata.normalize('NFKC', a or '')
    return re.sub(r'[\s　\*＊()（）]', '', a)


def surname(n):
    n = (n or '').strip()
    m = re.match(r'^([\u4e00-\u9fff]{1,2})', n)
    return m.group(1) if m else ''


def build_owner_keys(rows):
    """歸戶：同一遮罩統一編號可能對到多人，故以「統編＋姓氏＋地址群集」識別。
    地址一方為另一方前綴（謄本地址遮罩、門牌寫法不同）者視為同一人。"""
    buckets = {}
    for r in rows:
        buckets.setdefault((FL.norm_uid(r['統一編號']), surname(r['謄本姓名'])), []).append(r)
    for (u, sn), grp in buckets.items():
        addrs = sorted({norm_addr(r['通訊地址']) for r in grp if norm_addr(r['通訊地址'])},
                       key=len)
        rep = {}                                  # 地址 -> 代表地址
        for a in addrs:
            for b in addrs:
                if b is not a and len(b) <= len(a) and a.startswith(b):
                    rep[a] = rep.get(b, b); break
            else:
                rep.setdefault(a, a)
        for r in grp:
            a = norm_addr(r['通訊地址'])
            tag = rep.get(a, a) if a else ''
            r['地主鍵值'] = f'{u}|{sn}|{tag}' if tag else f'{u}|{sn}|~'


def style_header(ws, headers, manual=(), tracked=()):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = T_FILL if h in tracked else (M_FILL if h in manual else H_FILL)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'


def finish(ws, headers, widths=None, maxw=34):
    n = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=n, max_col=len(headers)):
        for c in row:
            c.border = THIN; c.font = Font(size=10)
            c.alignment = Alignment(vertical='top', wrap_text=True)
    for c, h in enumerate(headers, 1):
        if widths and h in widths:
            w = widths[h]
        else:
            w = max([len(str(h)) * 2] + [min(len(str(ws.cell(r, c).value or '')), 30) * 1.7
                                         for r in range(2, min(n, 300) + 1)] or [10])
            w = min(max(w + 2, 8), maxw)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(n, 1)}'


def dv(ws, headers, col, formula, rows=3000):
    if col not in headers: return
    d = DataValidation(type='list', formula1=formula, allow_blank=True)
    ws.add_data_validation(d)
    L = get_column_letter(headers.index(col) + 1)
    d.add(f'{L}2:{L}{rows}')


def main():
    src, out = sys.argv[1], sys.argv[2]
    block = sys.argv[3] if len(sys.argv) > 3 and not sys.argv[3].startswith('--') else 'A區'
    field_path = None
    if '--field' in sys.argv:
        field_path = sys.argv[sys.argv.index('--field') + 1]

    parcels = json.load(open(src))
    by_lot = {}
    for p in parcels:
        k = (p.get('段別'), p.get('地號全碼'))
        prev = by_lot.get(k)
        if prev is None or (p.get('謄本列印時間') or '') > (prev.get('謄本列印時間') or ''):
            by_lot[k] = p
    dropped = len(parcels) - len(by_lot)
    parcels = list(by_lot.values())
    if dropped: print(f'  同地號重複謄本 {dropped} 份，各保留列印時間最新者')

    # ── 現場名單比對
    fmap, only_field, only_deed, uid_issue = {}, [], [], []
    if field_path:
        frows = FL.load(field_path)
        deed_flat = [{'地號': p['地號'], '統編': FL.norm_uid(o.get('統一編號')),
                      '統編原文': o.get('統一編號'), '登記次序': o.get('登記次序'),
                      '姓名': o.get('所有權人')}
                     for p in parcels for o in p['所有權人清單']]
        fmap, only_field, only_deed, uid_issue = FL.match(frows, deed_flat)
        print(f'  現場名單 {len(frows)} 列：配對 {len(fmap)}、只在名單 {len(only_field)}、'
              f'只在謄本 {len(only_deed)}、統編疑義 {len(uid_issue)}')

    # ── 逐筆登記
    rows = []
    for p in parcels:
        area = p.get('面積_平方公尺') or 0
        cur = p.get('公告土地現值_元每平方公尺') or 0
        gid, gsize = joint_groups(p.get('所有權人清單', []))
        rights_by_seq = {}
        for r in p.get('他項權利清單', []):
            rights_by_seq.setdefault(r.get('標的登記次序', ''), []).append(
                f"{r.get('權利種類','')}／{r.get('權利人','')}")
        for o in p.get('所有權人清單', []):
            seq = o.get('登記次序', '')
            f = fmap.get((p['地號'], seq), {})
            fr = frac(o.get('權利範圍')); st = share_type(o.get('權利範圍'))
            if st == '公同共有' and fr is not None:
                n = gsize.get(gid.get(seq), 1) or 1
                fr = fr / n; st = f'公同共有(全體{n}人)'
            sh = float(area) * float(fr) if fr else None
            uid = o.get('統一編號', '')
            nm = f.get('名單姓名', '')
            rows.append({
                '區塊': block, '鄉鎮市': p.get('鄉鎮市', ''), '段別': p.get('段別', ''),
                '地號': p.get('地號', ''), '地號全碼': p.get('地號全碼', ''),
                '地目': f.get('地目', ''),
                '宗地面積(㎡)': area, '宗地面積(坪)': round(area / PING, 2) if area else None,
                '公告現值(元/㎡)': cur or None,
                '登記次序': seq,
                '地主代號': '', '全名': nm if FL.is_full_name(nm) else '',
                '謄本姓名': o.get('所有權人', ''), '統一編號': uid,
                '所有權人類型': owner_type(o.get('所有權人'), FL.norm_uid(uid)),
                '管理者': o.get('管理者', ''),
                '權利範圍': o.get('權利範圍', ''), '持分型態': st,
                '持分比例': round(float(fr), 6) if fr else None,
                '持分面積(㎡)': round(sh, 2) if sh is not None else None,
                '持分面積(坪)': round(sh / PING, 2) if sh is not None else None,
                '持分現值(元)': round(sh * cur) if sh is not None and cur else None,
                '電話': f.get('電話', ''), '通訊地址': f.get('通訊地址', ''),
                '意願等級': f.get('意願等級', ''), '承辦人': f.get('承辦人', ''),
                '登記日期': o.get('登記日期', ''), '登記原因': o.get('登記原因', ''),
                '權狀字號': o.get('權狀字號', ''),
                '設定他項權利': '；'.join(rights_by_seq.get(seq, [])),
                '其他登記事項': '／'.join(x for x in [
                    o.get('其他登記事項', ''),
                    ('地價備註：' + o['地價備註事項']) if o.get('地價備註事項') else '',
                ] if x),
                '同意書狀態': '', '紙本收件日': '', '印鑑證明': '未備', '聯絡方式': '',
                '上次接觸日': '', '下次追蹤日': '', '位置關鍵性': '', '備註': '',
                '謄本列印時間': re.sub(r'\s*頁次.*$', '', p.get('謄本列印時間', '')),
                '來源檔案': p.get('來源檔案', ''),
                '資料鍵值': f"{p['段別']}|{p['地號全碼']}|{seq}|{FL.norm_uid(uid)}",
                '_field': f,
            })

    # ── 讀既有檔案，保留人工資料與地主代號
    old_lot, old_own, old_visit, old_code, old_todo = {}, {}, [], {}, []
    if os.path.exists(out):
        wb0 = load_workbook(out)
        if '地號×地主明細' in wb0.sheetnames:
            ws0 = wb0['地號×地主明細']; h0 = [c.value for c in ws0[1]]
            for r in ws0.iter_rows(min_row=2, values_only=True):
                d = dict(zip(h0, r))
                if d.get('資料鍵值'): old_lot[d['資料鍵值']] = d
        if '拜訪名單（依地主）' in wb0.sheetnames:
            ws0 = wb0['拜訪名單（依地主）']; h0 = [c.value for c in ws0[1]]
            for r in ws0.iter_rows(min_row=2, values_only=True):
                d = dict(zip(h0, r))
                if d.get('地主鍵值'):
                    old_own[d['地主鍵值']] = d
                    if d.get('地主代號'): old_code[d['地主鍵值']] = d['地主代號']
        if '拜訪紀錄' in wb0.sheetnames:
            ws0 = wb0['拜訪紀錄']; h0 = [c.value for c in ws0[1]]
            for r in ws0.iter_rows(min_row=2, values_only=True):
                if any(v not in (None, '') for v in r):
                    old_visit.append(dict(zip(h0, r)))
        if '核對與待辦' in wb0.sheetnames:
            ws0 = wb0['核對與待辦']; h0 = [c.value for c in ws0[1]]
            for r in ws0.iter_rows(min_row=2, values_only=True):
                if any(v not in (None, '') for v in r):
                    old_todo.append(dict(zip(h0, r)))

    for r in rows:
        prev = old_lot.get(r['資料鍵值'])
        if prev:
            for m in M_LOT:
                if prev.get(m) not in (None, '', '未備'): r[m] = prev[m]
            for m in ('全名', '電話', '通訊地址', '意願等級', '承辦人', '地目'):
                if not r.get(m) and prev.get(m): r[m] = prev[m]

    build_owner_keys(rows)          # 地址回填完成後才歸戶

    # ── 地主層彙總
    own = {}
    for r in rows:
        k = r['地主鍵值']
        d = own.setdefault(k, {'謄本姓名': r['謄本姓名'], '全名': '', '統一編號': r['統一編號'],
                               '類型': r['所有權人類型'], '電話': '', '通訊地址': '',
                               '意願等級': '', '承辦人': '', '摘要': [], 'lots': [], 'n': 0,
                               'a': 0.0, 'v': 0})
        d['n'] += 1
        d['a'] += r.get('持分面積(㎡)') or 0
        d['v'] += r.get('持分現值(元)') or 0
        d['lots'].append(r['地號'])
        for f, key in (('全名', '全名'), ('電話', '電話'), ('通訊地址', '通訊地址'),
                       ('意願等級', '意願等級'), ('承辦人', '承辦人')):
            if not d[f] and r.get(key): d[f] = r[key]
        s = (r.get('_field') or {}).get('拜訪摘要', '')
        if s and s not in d['摘要']: d['摘要'].append(s.strip())

    total_area = sum(p.get('面積_平方公尺') or 0 for p in parcels)
    owners = []
    for k, d in own.items():
        lots = sorted(set(d['lots']), key=lambda s: (len(s), s))
        owners.append({
            '地主代號': old_code.get(k, ''), '全名': d['全名'], '謄本姓名': d['謄本姓名'],
            '統一編號': d['統一編號'], '電話': d['電話'], '通訊地址': d['通訊地址'],
            '所有權人類型': d['類型'], '持有地號': '、'.join(lots), '登記筆數': d['n'],
            '合計持分面積(㎡)': round(d['a'], 2), '合計持分面積(坪)': round(d['a'] / PING, 2),
            '佔已建檔比例': round(d['a'] / total_area, 4) if total_area else None,
            '意願等級': d['意願等級'], '拜訪階段': '', '承辦人': d['承辦人'],
            '關鍵人註記': '', '最近接觸日': '', '下次追蹤日': '', '同意書狀態': '',
            '紙本收件日': '', '印鑑證明': '未備',
            '最新拜訪摘要': ' ／ '.join(d['摘要']), '備註': '', '地主鍵值': k,
            '_摘要清單': list(d['摘要']),
        })
    owners.sort(key=lambda o: -o['合計持分面積(㎡)'])

    for o in owners:                                   # 併回既有人工欄位
        prev = old_own.get(o['地主鍵值'])
        if prev:
            for m in M_OWNER:
                if prev.get(m) not in (None, '', '未備'): o[m] = prev[m]
    nxt = 1
    used = {c for c in old_code.values() if c}
    for o in owners:
        if not o['地主代號']:
            while f'L{nxt:03d}' in used: nxt += 1
            o['地主代號'] = f'L{nxt:03d}'; used.add(o['地主代號']); nxt += 1
    code = {o['地主鍵值']: o['地主代號'] for o in owners}
    for i, o in enumerate(owners, 1): o['優先序'] = i
    for r in rows: r['地主代號'] = code.get(r['地主鍵值'], '')

    rows.sort(key=lambda r: (str(r['段別']), str(r['地號全碼']), str(r['登記次序'])))

    # ── 拜訪紀錄（從摘要拆解；已存在的人工列原樣保留）
    visits = list(old_visit)
    seen = {(str(v.get('日期', '')), str(v.get('地主代號', '')), str(v.get('內容摘要', ''))[:30])
            for v in visits}
    for o in owners:
        segs = [s2 for txt in o.get('_摘要清單', []) for s2 in FL.split_visits(txt)]
        for seg in segs:
            key = (str(seg['日期']), o['地主代號'], seg['內容'][:30])
            if key in seen: continue
            seen.add(key)
            visits.append({'日期': seg['日期'], '地主代號': o['地主代號'], '全名': o['全名'],
                           '謄本姓名': o['謄本姓名'], '地號': o['持有地號'],
                           '承辦人': o['承辦人'], '接觸方式': '', '對象': '',
                           '內容摘要': seg['內容'], '意願等級': o['意願等級'],
                           '下次行動': '', '下次日期': '', '登錄者': ''})
    visits.sort(key=lambda v: (str(v.get('日期') or '9999'), str(v.get('地主代號') or '')))

    # ── 他項權利
    rights, seenr = [], set()
    for p in parcels:
        for r in p.get('他項權利清單', []):
            x = [block, p.get('地號', ''), p.get('地號全碼', ''), r.get('序號', ''),
                 r.get('登記次序', ''), r.get('權利種類', ''), r.get('權利人', ''),
                 r.get('權利人統一編號', ''), r.get('擔保債權總金額'),
                 r.get('標的登記次序', ''), r.get('設定權利範圍', ''), r.get('存續期間', ''),
                 r.get('共同擔保地號', ''), r.get('登記日期', ''), r.get('登記原因', '')]
            k = tuple(map(str, x[1:6]))
            if k not in seenr: seenr.add(k); rights.append(x)
    rights.sort(key=lambda x: (len(str(x[2])), str(x[2]), str(x[3])))

    # ═══ 產出 ═══
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet('拜訪名單（依地主）')
    style_header(ws, H_OWNER, M_OWNER)
    for o in owners: ws.append([o.get(h) for h in H_OWNER])
    for h, fmt in [('合計持分面積(㎡)', '#,##0.00'), ('合計持分面積(坪)', '#,##0.00'),
                   ('佔已建檔比例', '0.00%')]:
        ci = H_OWNER.index(h) + 1
        for r in range(2, ws.max_row + 1): ws.cell(r, ci).number_format = fmt
    for col, f in [('意願等級', WILL), ('拜訪階段', STAGE), ('印鑑證明', SEAL)]:
        dv(ws, H_OWNER, col, f)
    finish(ws, H_OWNER, {'最新拜訪摘要': 46, '通訊地址': 30, '持有地號': 22, '備註': 24})
    ws.column_dimensions[get_column_letter(H_OWNER.index('地主鍵值') + 1)].hidden = True

    ws2 = wb.create_sheet('地號×地主明細')
    style_header(ws2, H_LOT, M_LOT)
    for r in rows: ws2.append([r.get(h) for h in H_LOT])
    for h, fmt in [('宗地面積(㎡)', '#,##0.00'), ('宗地面積(坪)', '#,##0.00'),
                   ('公告現值(元/㎡)', '#,##0'), ('持分比例', '0.000000'),
                   ('持分面積(㎡)', '#,##0.00'), ('持分面積(坪)', '#,##0.00'),
                   ('持分現值(元)', '#,##0')]:
        ci = H_LOT.index(h) + 1
        for r in range(2, ws2.max_row + 1): ws2.cell(r, ci).number_format = fmt
    for col, f in [('同意書狀態', STAGE), ('印鑑證明', SEAL), ('位置關鍵性', KEY),
                   ('意願等級', WILL)]:
        dv(ws2, H_LOT, col, f)
    finish(ws2, H_LOT, {'其他登記事項': 30, '通訊地址': 28})
    ws2.column_dimensions[get_column_letter(H_LOT.index('資料鍵值') + 1)].hidden = True

    ws3 = wb.create_sheet('地號彙總')
    style_header(ws3, H_SUM, ('位置關鍵性', '備註'))
    par = {}
    for r in rows:
        k = (r['段別'], r['地號全碼'])
        d = par.setdefault(k, {'r': r, 'n': 0, 'sum': Fraction(0), 'types': set(),
                               'jt': False, 'joint': set(), 'A': 0})
        d['n'] += 1
        st = r.get('持分型態', '') or ''
        fr = frac(r.get('權利範圍'))
        if st.startswith('公同共有'):
            d['jt'] = True
            gk = (r['段別'], r['地號全碼'], r.get('權利範圍'))
            if gk not in d['joint']:
                d['joint'].add(gk)
                if fr: d['sum'] += fr
        elif fr:
            d['sum'] += fr
        d['types'].add(r.get('所有權人類型', ''))
        if str(r.get('意願等級', '')).upper() == 'A': d['A'] += 1
    rcount = {}
    for x in rights: rcount[str(x[2])] = rcount.get(str(x[2]), 0) + 1
    rn = 1
    for (seg, no), d in sorted(par.items(), key=lambda kv: (len(kv[0][1]), kv[0][1])):
        r = d['r']; rn += 1; s = float(d['sum'])
        pub = [t for t in d['types'] if t.startswith('公有')]
        ws3.append([r['區塊'], seg, r['地號'], no, r['地目'], r['宗地面積(㎡)'],
                    r['宗地面積(坪)'], r['公告現值(元/㎡)'], d['n'],
                    '公同共有' if d['jt'] else ('單獨所有' if d['n'] == 1 else '分別共有'),
                    '✔ 1/1' if abs(s - 1) < 1e-9 else f'⚠ {s:.4f}',
                    '公有' if pub and len(d['types']) == 1 else ('公私共有' if pub else '私有'),
                    rcount.get(no, 0), d['A'],
                    f'=IF($I{rn}=0,"",TEXT($N{rn}/$I{rn},"0%"))', '', ''])
    for h, fmt in [('宗地面積(㎡)', '#,##0.00'), ('宗地面積(坪)', '#,##0.00'),
                   ('公告現值(元/㎡)', '#,##0')]:
        ci = H_SUM.index(h) + 1
        for r in range(2, ws3.max_row + 1): ws3.cell(r, ci).number_format = fmt
    dv(ws3, H_SUM, '位置關鍵性', KEY, 500)
    finish(ws3, H_SUM)

    ws4 = wb.create_sheet('拜訪紀錄')
    style_header(ws4, H_VISIT, tracked=H_VISIT)
    for v in visits: ws4.append([v.get(h) for h in H_VISIT])
    for col, f in [('意願等級', WILL),
                   ('接觸方式', '"面訪,電話,LINE,簡訊,他人轉達,說明會"'),
                   ('對象', '"本人,配偶,子女,父母,親屬,代理人,鄰居,里長,未接觸"')]:
        dv(ws4, H_VISIT, col, f)
    finish(ws4, H_VISIT, {'內容摘要': 60, '地號': 20})

    ws5 = wb.create_sheet('他項權利')
    style_header(ws5, H_RIGHT)
    for x in rights: ws5.append(x)
    ci = H_RIGHT.index('擔保債權總金額') + 1
    for r in range(2, ws5.max_row + 1): ws5.cell(r, ci).number_format = '#,##0'
    finish(ws5, H_RIGHT)

    ws6 = wb.create_sheet('核對與待辦')
    style_header(ws6, H_TODO, ('負責人', '完成日'))
    done = {(str(t.get('類別')), str(t.get('地號')), str(t.get('對象'))):
            (t.get('負責人'), t.get('完成日')) for t in old_todo}

    def todo(cat, lot, who, uid, desc, act):
        p_ = done.get((str(cat), str(lot), str(who)), ('', ''))
        ws6.append([cat, lot, who, uid, desc, act, p_[0] or '', p_[1] or ''])

    if not field_path and old_todo:
        # 本次未帶現場名單，比對結果無法重算 → 沿用既有待辦
        for t in old_todo:
            ws6.append([t.get(h) for h in H_TODO])
    else:
        deed_lots = {p['地號'] for p in parcels}
        field_lots = sorted({f['地號'] for f in (FL.load(field_path) if field_path else [])},
                            key=lambda s: (len(s), s))
        for lot in field_lots:
            if lot not in deed_lots:
                todo('待調謄本', lot, '', '', '名單有此地號，但尚未調閱土地登記謄本',
                     '調第二類謄本後併入名冊')
        for d in only_deed:
            todo('名單漏列地主', d['地號'], d.get('姓名', ''), d.get('統編原文', ''),
                 f"謄本有此所有權人（登記次序 {d.get('登記次序','')}），現場名單未列",
                 '補入拜訪名單並安排拜訪')
        for f in only_field:
            todo('名單多出地主', f['地號'], f.get('名單姓名', ''), f.get('統編原文', ''),
                 '名單有此人，但該地號謄本查無（多為尚未調謄本之地號）', '調謄本後複核')
        for u in uid_issue:
            todo('統一編號疑義', u['地號'], u['名單姓名'], u['名單統編'],
                 f"名單 {u['名單統編']} vs 謄本 {u['謄本統編']}：首碼不同、數字相同",
                 '以謄本為準，修正名單統編')
        dupes = {}
        for o in owners:
            dupes.setdefault(FL.norm_uid(o['統一編號']), []).append(o)
        for u, lst in dupes.items():
            if len(lst) > 1 and u:
                todo('遮罩統編一碼多人', '',
                     '／'.join(x['全名'] or x['謄本姓名'] for x in lst), lst[0]['統一編號'],
                     f'同一遮罩統一編號對到 {len(lst)} 位不同地主（已依通訊地址分開歸戶）',
                     '確認地址無誤；必要時調第一類謄本')
    finish(ws6, H_TODO, {'說明': 44, '建議處理': 26, '對象': 26})

    # ── 說明與進度
    ws7 = wb.create_sheet('說明與進度')
    n_lots, n_rows_, n_owners = len(par), len(rows), len(owners)
    aA = sum(o['合計持分面積(㎡)'] for o in owners if str(o['意願等級']).upper() == 'A')
    nA = sum(1 for o in owners if str(o['意願等級']).upper() == 'A')
    n_phone = sum(1 for o in owners if o['電話'])
    n_name = sum(1 for o in owners if o['全名'])
    A = [
        ('員林交流道附近特定區｜產業發展儲備用地　地主同意書整合名冊（出勤版）', ''),
        ('', ''),
        ('■ 怎麼用', ''),
        ('出門前', '印「拜訪名單（依地主）」，已依合計持分面積排優先序；一人一列，一份同意書'),
        ('回來後', '每次接觸在「拜訪紀錄」新增一列（日期／承辦人／接觸方式／內容），'
                  '再回「拜訪名單」更新意願等級與拜訪階段'),
        ('橘色標題欄位', '人工填寫；深藍色為謄本自動帶入，請勿手改（下批併檔會覆蓋）'),
        ('棕色標題工作表', '「拜訪紀錄」整張都是人工登錄，併檔不會動它'),
        ('', ''),
        ('■ 意願等級定義', ''),
        ('A', '明確同意／高意願，可安排簽同意書'),
        ('B', '可談，尚需確認（等說明會、要問家人、共有人同意就同意）'),
        ('C', '觀望，態度未明'),
        ('D', '明確反對／婉拒'),
        ('E', '聯絡不上（查無地址、電話不通、長期不在）'),
        ('未接觸', '尚未拜訪'),
        ('', ''),
        ('■ 目前彙總', ''),
        ('已建檔地號筆數', n_lots),
        ('已建檔面積（㎡）', round(total_area, 2)),
        ('已建檔面積（坪）', round(total_area / PING, 2)),
        ('已建檔面積（公頃）', round(total_area / 10000, 4)),
        ('所有權登記筆數', n_rows_),
        ('不重複地主人數（＝同意書份數）', n_owners),
        ('已取得全名', f'{n_name} 人（{n_name/n_owners*100:.0f}%）' if n_owners else ''),
        ('已取得電話', f'{n_phone} 人（{n_phone/n_owners*100:.0f}%）' if n_owners else ''),
        ('意願等級 A 人數', f'{nA} 人，合計持分 {aA:,.0f} ㎡'
                          f'（佔已建檔 {aA/total_area*100:.1f}%）' if total_area else nA),
        ('他項權利（抵押權等）筆數', len(rights)),
        ('待辦事項筆數', ws6.max_row - 1),
        ('', ''),
        ('■ 門檻對照（來源：主要計畫四通 表9-1，頁9-2）', ''),
        ('A區公告面積', '13.18 公頃 = 131,800 ㎡'),
        ('變更面積門檻', '不得小於 5 公頃（都市計畫農業區變更使用審議規範 §12）'),
        ('已建檔面積佔 A 區比例', f'{total_area/131800*100:.2f}%'),
        ('已建檔面積佔 5 公頃門檻比例', f'{total_area/50000*100:.2f}%'),
        ('', ''),
        ('■ 同意門檻（兩階段，不可混用）', ''),
        ('變更階段（現階段）', '申請範圍內全體土地所有權人同意（100%），須檢具土地使用同意書'),
        ('重劃階段（變更發布後）', '私有土地所有權人半數以上且面積超過半數（平均地權條例 §58）'),
        ('送件硬期限', '117.08.31（主要計畫 112.08.31 發布實施起 5 年內）'),
        ('', ''),
        ('■ 工作鐵律', ''),
        ('1', '口頭應允、LINE 回訊一律歸「B」，只有紙本簽回才計入同意率'),
        ('2', '催收優先序依「位置關鍵性」排，不依面積大小排（卡道路／卡中央的小坪數殺傷力最大）'),
        ('3', '公同共有須共有人全體同意（民法 §828），缺一不可'),
        ('4', '遮罩統一編號一碼可能對到多人，本檔以「統編＋通訊地址」歸戶；'
              '地址不同即視為不同人，勿再以統編合併'),
        ('5', '面積為謄本登載值，實際面積以核定圖實地分割測量面積為準'),
        ('6', '本檔含個資，僅限內部使用；不得外流、不得寫入對外文案或公開連結'),
        ('', ''),
        ('■ 資料來源與產製', ''),
        ('謄本', '土地登記第二類謄本（權利範圍、面積、他項權利以此為準）'),
        ('現場名單', os.path.basename(field_path) if field_path else '（本次未帶入）'),
        ('產製日期', datetime.date.today().isoformat()),
        ('後續批次', '再上傳謄本 PDF 即可併入本檔，人工填寫欄位與拜訪紀錄會自動保留'),
    ]
    ws7.column_dimensions['A'].width = 34; ws7.column_dimensions['B'].width = 88
    for i, (a, b) in enumerate(A, 1):
        ws7.cell(i, 1, a); ws7.cell(i, 2, b)
        if str(a).startswith('■') or i == 1:
            ws7.cell(i, 1).font = Font(bold=True, size=12 if i == 1 else 11, color='1F3864')
        ws7.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='center')
    wb.move_sheet('說明與進度', offset=-6)

    wb.save(out)
    print(f'OK -> {out}')
    print(f'地號 {n_lots}｜登記 {n_rows_} 筆｜地主 {n_owners} 人｜面積 {total_area:,.2f} ㎡'
          f' ({total_area/10000:.4f} ha)｜意願A {nA} 人｜拜訪紀錄 {len(visits)} 筆'
          f'｜待辦 {ws6.max_row-1} 筆')


main()
