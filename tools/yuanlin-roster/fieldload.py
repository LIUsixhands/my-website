# -*- coding: utf-8 -*-
"""讀取現場跑地主名單（含全名/電話/地址/意願等級/拜訪狀況/承辦人），
以「地號 × 統一編號」對回謄本資料；首字母不同但數字相同者做模糊比對並標記。"""
import re, unicodedata
from collections import defaultdict
from openpyxl import load_workbook

def norm_lot(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'): s = s[:-2]
    s = unicodedata.normalize('NFKC', s).replace('–', '-').replace('—', '-')
    m = re.match(r'^(\d+)(?:-(\d+))?$', s)
    if m:
        return str(int(m.group(1))) + ('-' + str(int(m.group(2))) if m.group(2) else '')
    return s

def norm_uid(v):
    if v is None: return ''
    s = str(v).strip()
    if s.endswith('.0'): s = s[:-2]
    s = unicodedata.normalize('NFKC', s).upper()
    if s == '158': return '0000000158'
    return re.sub(r'[*\s\-　]', '', s)

def uid_digits(u):
    return re.sub(r'^[A-Z]', '', u)          # 去掉首碼英文字母

def clean(v):
    if v is None: return ''
    if isinstance(v, str): return v.strip()
    s = str(v)
    return s[:-2] if s.endswith('.0') else s

def norm_phone(v):
    s = clean(v)
    if not s: return ''
    out = []
    for part in re.split(r'[\n/、,，]+', s):
        p = part.strip()
        if not p: continue
        d = re.sub(r'\D', '', p)
        if d and not re.search(r'[一-鿿]', p):
            if len(d) == 9 and d.startswith('9'): d = '0' + d      # 912635425 -> 0912635425
            p = d
        out.append(p)
    return '／'.join(out)

def is_full_name(n):
    """判斷是否為真實全名（非「陳先生」「陳＊＊」這類代稱）"""
    n = (n or '').strip()
    if not n or len(n) < 2: return False
    if '＊' in n or '*' in n: return False
    if re.search(r'(先生|小姐|女士|太太)$', n): return False
    if n in ('中華民國',): return False
    return bool(re.fullmatch(r'[一-鿿]{2,5}', n))

def load(path, sheet=None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        r = list(r) + [None] * 20
        if r[2] is None: continue
        rows.append({
            '來源列': i, '地號': norm_lot(r[2]), '地目': clean(r[5]),
            '名單姓名': clean(r[7]), '統編': norm_uid(r[8]), '統編原文': clean(r[8]),
            '登記日期': clean(r[9]), '登記原因': clean(r[10]),
            '通訊地址': clean(r[11]), '他項摘要': clean(r[12]),
            '電話': norm_phone(r[13]), '意願等級': clean(r[14]),
            '拜訪摘要': clean(r[15]), '承辦人': clean(r[16]),
        })
    return rows

def match(field_rows, deed_rows):
    """回傳 (配對表, 只在名單, 只在謄本, 統編疑義)
    deed_rows: [{'地號','統編','登記次序',...}]"""
    fk, dk = defaultdict(list), defaultdict(list)
    for f in field_rows:
        if f['統編']: fk[(f['地號'], f['統編'])].append(f)
    for d in deed_rows: dk[(d['地號'], d['統編'])].append(d)

    pair, uid_issue = {}, []
    used_f, used_d = set(), set()
    for k in set(fk) & set(dk):                      # 完全相符
        for f, d in zip(fk[k], dk[k]):
            pair[(d['地號'], d['登記次序'])] = f
            used_f.add(id(f)); used_d.add((d['地號'], d['登記次序']))
        for f in fk[k][len(dk[k]):]: used_f.add(id(f))
        for d in dk[k][len(fk[k]):]: used_d.add((d['地號'], d['登記次序']))

    # 模糊：同地號、統編數字相同但首碼字母不同（名單常見打字差異）
    rest_f = defaultdict(list); rest_d = defaultdict(list)
    for f in field_rows:
        if id(f) not in used_f and f['統編']:
            rest_f[(f['地號'], uid_digits(f['統編']))].append(f)
    for d in deed_rows:
        if (d['地號'], d['登記次序']) not in used_d:
            rest_d[(d['地號'], uid_digits(d['統編']))].append(d)
    for k in set(rest_f) & set(rest_d):
        for f, d in zip(rest_f[k], rest_d[k]):
            pair[(d['地號'], d['登記次序'])] = f
            used_f.add(id(f)); used_d.add((d['地號'], d['登記次序']))
            uid_issue.append({'地號': d['地號'], '登記次序': d['登記次序'],
                              '名單姓名': f['名單姓名'],
                              '名單統編': f['統編原文'], '謄本統編': d.get('統編原文', ''),
                              '說明': '統編首碼不同、數字相同，已依地號比對；謄本為準，建議修正名單'})

    only_field = [f for f in field_rows if id(f) not in used_f]
    only_deed = [d for d in deed_rows if (d['地號'], d['登記次序']) not in used_d]
    return pair, only_field, only_deed, uid_issue

VISIT_DATE = re.compile(r'(\d{1,2})\s*/\s*(\d{1,2})')

def split_visits(text, year=2026):
    """把「8/24…8/26…」的自由文字拆成逐次拜訪紀錄"""
    t = (text or '').replace('\n', ' ').strip()
    if not t: return []
    ms = list(VISIT_DATE.finditer(t))
    if not ms: return [{'日期': '', '內容': t}]
    out = []
    if ms[0].start() > 0:
        head = t[:ms[0].start()].strip(' ，,、')
        if head: out.append({'日期': '', '內容': head})
    for i, m in enumerate(ms):
        end = ms[i + 1].start() if i + 1 < len(ms) else len(t)
        body = t[m.end():end].strip(' ，,、')
        try:
            d = f"{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        except ValueError:
            d = m.group(0)
        out.append({'日期': d, '內容': body or '(未記述)'})
    return out
